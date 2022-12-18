from openpyxl import load_workbook
def get():
    filename = 'my_project.xlsx'
    wb = load_workbook(filename)
    sheet = wb.active
    id_ = 1 
    surname = input('Введите фамилию: ')
    first_name = input('Введите имя: ')
    patronym = input('Введите отчество: ')
    phone = input('Введите телефон: ')
    comment = input('Введите комментарий: ')
    for row in range(sheet.max_row + 1, sheet.max_row + 2):        
        sheet[row][0].value = row - 1
        sheet[row][1].value = surname 
        sheet[row][2].value = first_name 
        sheet[row][3].value = patronym 
        sheet[row][4].value = phone
        sheet[row][5].value = comment
    wb.save(filename) 
    wb.close() 
